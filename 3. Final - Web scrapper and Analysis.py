#WEBSCRAPER AND EXCEL EXPORT FILES

# Should the user want to change the input data, delete the old info on input file (Micasa) and re-run webscrapper and analysis

# Check if analysis was run before and prepare the file for webscrapper
try: 
    
    import getpass
    username = getpass.getuser()
    from openpyxl import load_workbook
    worksheet = load_workbook("C:/Users/%s/Desktop/Micasaestucasa/working file/Micasa.xlsx" %username)
    worksheet.remove(worksheet.get_sheet_by_name("Descriptors"))
    worksheet.remove(worksheet.get_sheet_by_name("Listings per room"))
    worksheet.remove(worksheet.get_sheet_by_name("Average price per neighbourhood")) 
    worksheet.save("C:/Users/%s/Desktop/Micasaestucasa/working file/Micasa.xlsx" %username)

except:

    worksheet.save("C:/Users/%s/Desktop/Micasaestucasa/working file/Micasa.xlsx" %username)

# web scrapper takes the input from the Micasa excel and downloads the CSV on the Insideairbnb website
import pandas as pd

# Read user input for the web scrapper
wb = load_workbook(filename = (r'C:/Users/%s/Desktop/Micasaestucasa/Working File/\Micasa.xlsx' %username))
sheet = wb["Micasa"]
minput = sheet['B2'].value
yinput = sheet['B3'].value
pinput = sheet['B4'].value
monthinput = minput
yearinput = yinput
placeinput = pinput

# Display the info (in the python console) that will be downloaded as a sanity check for the user
print(monthinput)
print(yearinput)
print(placeinput)

# import librairies needed for web scrapping
import requests 
import re
from bs4 import BeautifulSoup as bs

# Load the webpage content
r = requests.get("http://insideairbnb.com/get-the-data.html")

# Convert to a beautiful soup object
webpage = bs(r.content, "html.parser") 

# Print out our html to be exported as a check for the user
for link in webpage.findAll('a', attrs={'href': re.compile(".%s/%s.%s-.*/visualisations/listings.csv$" % (placeinput, yearinput, monthinput))}):
    print(link.get('href'))
  
# Import the file into the right folder on the computer
df = pd.read_csv(link.get('href'))
df.head()
df.to_csv(r'C:/Users/%s/Desktop/Micasaestucasa/export/\out.csv' %username)

# Rename the new import based on the user input
import os
os.rename(r'C:/Users/%s/Desktop/Micasaestucasa/export/\out.csv' %username, (r'C:/Users/%s/Desktop/Micasaestucasa/export/\%s %s.%s.csv' %(username, placeinput, monthinput, yearinput)))
read_file = pd.read_csv (r'C:\Users\%s\Desktop\Micasaestucasa\export\%s %s.%s.csv' %(username, placeinput, monthinput, yearinput))
read_file.to_excel (r'C:\Users\%s\Desktop\Micasaestucasa\export\%s %s.%s.xlsx' %(username, placeinput, monthinput, yearinput), index = None, header=True)




#------------------------------------------------------------------------------------------------

# This code section builds the necessary tabs on the Micasa file for our analysis

# Read the data that we specifically need from our export to speed the analysis processing time
df = pd.read_excel(r'C:/Users/%s/Desktop/Micasaestucasa/export/%s %s.%s.xlsx' %(username, placeinput, monthinput, yearinput), usecols='D, G, J, K')

# Show a sample in the python console as a sanity check (first and last) and describe the dataset
print (df.head())
print(df.tail())
s = df.describe()

# create the necessary tabs for analysis
worksheet = load_workbook("C:/Users/%s/Desktop/Micasaestucasa/working file/Micasa.xlsx" %username)
worksheet.create_sheet("Descriptors") 
worksheet.create_sheet("Listings per room") 
worksheet.create_sheet("Average price per neighbourhood") 
worksheet.save("C:/Users/%s/Desktop/Micasaestucasa/working file/Micasa.xlsx" %username)

# print the description of the data set in the descriptors sheet of the Micasa file
import xlwings as xw
wb = xw.Book("C:/Users/%s/Desktop/Micasaestucasa/working file/Micasa.xlsx" %username)
sheet = wb.sheets("Descriptors") 
sheet.range('A1').value = s
wb.save("C:/Users/%s/Desktop/Micasaestucasa/working file/Micasa.xlsx" %username)


#---------------------------------------------

# Import necessary librairies for data analysis and visualizations
import matplotlib.pyplot as plt
import seaborn as sns
import pandas as pd


# This section is our first analysis of the extracted data (analysis of the avg price per room type)
# Read the necessary data in our export and plot as a bar chart in the 'Listing per room' sheet created earlier
listings_data = pd.read_excel(r'C:/Users/%s/Desktop/Micasaestucasa/export/%s %s.%s.xlsx' %(username, placeinput, monthinput, yearinput),
              usecols='D, G, J, K')
listings = listings_data[['host_id','neighbourhood','room_type','price']]
sns.set(style="whitegrid")
fig, ax = plt.subplots()
fig.set_size_inches(8, 5)
ax.axes.set_title("Number of listings per room type",fontsize=20, pad=20)
ax = sns.countplot(y='room_type',data=listings,order=listings['room_type'].value_counts().index, palette="Set3")
ax.set_xlabel('Room type',fontsize=14,labelpad=15)
ax.set_ylabel('Type of Room',fontsize=14,labelpad=15)
ax.xaxis.set_tick_params(labelsize=10)
ax.yaxis.set_tick_params(labelsize=10)

# Print the analysis (graph) in the excel and save the excel
sheet2 = wb.sheets("Listings per room")
rng = wb.sheets("Listings per room").range("B2")
sheet2.pictures.add(fig, top=rng.top, left=rng.left)

wb.save("C:/Users/%s/Desktop/Micasaestucasa/working file/Micasa.xlsx" %username)

#---------------------------------------------

# This section is our second analysis of the extracted data (analysis of the avg price of listings per neighbourhood)
top = listings['neighbourhood'].value_counts().iloc[:20].index.tolist()

# Create a dataframe to group neighbourhoods by average room price
price = pd.DataFrame(listings.groupby(['neighbourhood']).price.mean().reset_index())

# Create a dataframe to filter top 20 neighbourhoods
top_price = price[price['neighbourhood'].isin(top)].sort_values('price',ascending=False)

# Barplot of the avg price by neighbourhood, only display the top 20 neighbourhoods in terms of listings
fig, ax = plt.subplots()
fig.set_size_inches(8, 5)
ax.axes.set_title("Room Price",fontsize=20, pad=20)
ax = sns.barplot(x='price', y='neighbourhood',data=top_price, palette='Set3')
ax.set_xlabel('Avg. Price',fontsize=20,labelpad=20)
ax.set_ylabel('Neighbourhood',fontsize=20,labelpad=20)
ax.xaxis.set_tick_params(labelsize=10)
ax.yaxis.set_tick_params(labelsize=10)

# Print the analysis (graph) in the excel and save the excel
sheet3 = wb.sheets("Average price per neighbourhood")
rng = wb.sheets("Average price per neighbourhood").range("B2")
sheet3.pictures.add(fig, top=rng.top, left=rng.left)

wb.save("C:/Users/%s/Desktop/Micasaestucasa/working file/Micasa.xlsx" %username)

# Make a copy of the excel and rename it to display the month, date year of the analysis performed
import shutil

original = r'C:/Users/%s/Desktop/Micasaestucasa/working file/Micasa.xlsx' %username
target = r'C:/Users/%s/Desktop/Micasaestucasa/working file/Micasa %s %s.%s.xlsx' %(username, placeinput, monthinput, yearinput)

shutil.copyfile(original, target)